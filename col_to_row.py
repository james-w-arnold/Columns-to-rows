import openpyxl
import argparse
def main():
    parser = argparse.ArgumentParser(description="Defining the input and output files")
    parser.add_argument('-i', '--input', type=str, help="Input file")
    parser.add_argument('-o', '--output', type=str, help="location of output")
    parser.add_argument('-s', '--sheet', type=str, help="Name of the sheet on the spreadsheet")
    parser.add_argument('-c', '--startcell', type=str, help="Location of the start cell (in two digits, so C5 would be (3,5))")
    parser.add_argument('-e', '--endcell', type=str, help="location of the last cell (in two digits)")
    args = parser.parse_args()
    wb = openpyxl.load_workbook(args.input, data_only=True)
    print(args)
    reformatted_range = []
    sheet = wb.get_sheet_by_name(args.sheet)

    start_cell_column, start_cell_row = args.startcell.split(',')
    start_cell_column = int(start_cell_column)
    start_cell_row = int(start_cell_row)

    end_cell_column, end_cell_row = args.endcell.split(',')
    end_cell_column = int(end_cell_column)
    end_cell_row = int(end_cell_row)

    for i in range(start_cell_row, end_cell_row+1):
        #print(sheet.cell(row=i, column=19).value)
        for j in range(start_cell_column+1, end_cell_column+1):
            cell = sheet.cell(row=i, column=j).value
            if cell is not None and cell != "#VALUE!" and cell is not "0" and cell != "#REF!" and cell is not 0 and cell is not 5 and cell is not 50:
                reformatted_range.append((str(sheet.cell(row=i, column=start_cell_column).value), str(cell)))
            #print(sheet.cell(row=i, column=j).value)
    new_book = openpyxl.Workbook()
    new_sheet = new_book.active
    for i in range(1, len(reformatted_range)):
        a = "A{}".format(i)
        b = "B{}".format(i)
        new_sheet[a] = reformatted_range[i-1][0]
        new_sheet[b] = reformatted_range[i-1][1]

    new_book.save(args.output)

if __name__ == '__main__':
    main()
